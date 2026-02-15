"""
Script to evaluate all peptide models across three GNN layer configurations,
populate Excel table with results (using MEDIAN), and highlight best values.
"""

import glob
import pickle
import numpy as np
import pandas as pd
import math
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy.stats import friedmanchisquare
import scikit_posthocs as sp

# TensorFlow/Keras imports
from tensorflow.keras.models import load_model
from stellargraph.layer import DeepGraphCNN, GCNSupervisedGraphClassification, SortPooling, GraphConvolution
from stellargraph.mapper import PaddedGraphGenerator

# Sklearn metrics
from sklearn.metrics import matthews_corrcoef, precision_score, recall_score, f1_score, roc_auc_score, confusion_matrix

# Add parent directory for imports
sys.path.append(os.path.abspath(os.path.dirname(__file__)))
from cv_splits.data_splits import load_cv_splits


# ============== METRIC FUNCTIONS ==============
def roc_auc_metric(y_true, y_pred):
    return roc_auc_score(y_true, y_pred)

def rest_of_metrics(y_true, y_pred):
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred).ravel()
    tpr = tp / (tp + fn + 1e-8)
    tnr = tn / (tn + fp + 1e-8)
    gm = math.sqrt(tpr * tnr)
    precision = precision_score(y_true, y_pred, zero_division=0)
    recall = recall_score(y_true, y_pred, zero_division=0)
    f1 = f1_score(y_true, y_pred, zero_division=0)
    return gm, precision, recall, f1

def mcc_metric(y_true, y_pred):
    return matthews_corrcoef(y_true, y_pred)

def extract_fold_number(filename):
    return int(filename.split("_fold_")[-1].split(".")[0])

def load_models_by_pattern(pattern):
    model_paths = glob.glob(pattern)
    models = []
    for path in model_paths:
        model = load_model(
            path,
            custom_objects={
                "DeepGraphCNN": DeepGraphCNN,
                "GCNSupervisedGraphClassification": GCNSupervisedGraphClassification,
                "SortPooling": SortPooling,
                "GraphConvolution": GraphConvolution,
            }
        )
        models.append((os.path.basename(path), model))
    return models


# ============== EVALUATION FUNCTION ==============
def evaluate_configuration(config_name, processed_file, splits_file, model_patterns):
    """
    Evaluate a single GNN layer configuration.
    
    Args:
        config_name: Name of the configuration (for logging)
        processed_file: Path to the processed data pickle file
        splits_file: Path to the CV splits pickle file
        model_patterns: Dict with keys 'baseline', 'method1', 'method2', 'method3', 'method4'
                       containing glob patterns for model files
    
    Returns:
        metrics_summary: Dict with median/std for each model group and metric
        friedman_pvalues: Dict with Friedman p-values for each metric
        significant_vs_baseline: Dict indicating if best model is significantly better than baseline
    """
    print(f"\n{'='*60}")
    print(f"Evaluating configuration: {config_name}")
    print(f"{'='*60}")
    
    # Load data
    with open(processed_file, "rb") as f:
        processed_data = pickle.load(f)
    
    graphs = np.array(processed_data["graphs"])
    labels = np.array(processed_data["labels"])
    splits = load_cv_splits(splits_file)
    
    # Load models
    model_groups_loaded = {}
    for group_name, pattern in model_patterns.items():
        models = load_models_by_pattern(pattern)
        if len(models) == 0:
            print(f"  WARNING: No models found for {group_name} with pattern {pattern}")
        else:
            print(f"  Loaded {len(models)} models for {group_name}")
        model_groups_loaded[group_name] = sorted(models, key=lambda x: extract_fold_number(x[0]))
    
    gen = PaddedGraphGenerator(graphs=graphs)
    
    # Evaluate each fold
    results = {}
    for fold_idx, fold in enumerate(splits, start=1):
        test_idx = fold["test_idx"]
        X_test = graphs[test_idx]
        y_test = labels[test_idx]
        test_gen = gen.flow(X_test, y_test, batch_size=32, shuffle=False)
        
        results[fold_idx] = {}
        
        for group_name, models_list in model_groups_loaded.items():
            if len(models_list) < fold_idx:
                continue
            filename, model = models_list[fold_idx - 1]
            
            y_pred_probs = model.predict(test_gen, verbose=0)
            y_pred_probs = np.reshape(y_pred_probs, (-1,))
            
            roc_auc = roc_auc_metric(y_test, y_pred_probs)
            y_pred = (y_pred_probs >= 0.5).astype(int)
            gm, precision, recall, f1 = rest_of_metrics(y_test, y_pred)
            mcc = mcc_metric(y_test, y_pred)
            
            results[fold_idx][group_name] = {
                "ROC_AUC": roc_auc,
                "GM": gm,
                "Precision": precision,
                "Recall": recall,
                "F1": f1,
                "MCC": mcc,
            }
    
    # Compute summary statistics
    model_groups = ["baseline", "method1", "method2", "method3", "method4"]
    metric_names = ["ROC_AUC", "GM", "Precision", "Recall", "F1", "MCC"]
    
    metrics_summary = {}
    fold_results = {group: {metric: [] for metric in metric_names} for group in model_groups}
    
    for fold in sorted(results.keys()):
        for group in model_groups:
            if group in results[fold]:
                for metric in metric_names:
                    fold_results[group][metric].append(results[fold][group][metric])
    
    for group in model_groups:
        metrics_summary[group] = {}
        for metric in metric_names:
            values = fold_results[group][metric]
            if len(values) > 0:
                metrics_summary[group][metric] = {
                    "median": np.median(values),
                    "std": np.std(values)
                }
            else:
                metrics_summary[group][metric] = {"median": np.nan, "std": np.nan}
    
    # Friedman test and significance
    friedman_pvalues = {}
    best_significant_vs_baseline = {}  # Only tracks if BEST model is significant vs baseline
    
    for metric in metric_names:
        data_for_metric = {model: fold_results[model][metric] for model in model_groups}
        df_metric = pd.DataFrame(data_for_metric)
        
        # Check if we have enough data
        if df_metric.dropna().shape[0] < 3:
            friedman_pvalues[metric] = np.nan
            best_significant_vs_baseline[metric] = False
            continue
        
        data_list = [df_metric[col].dropna().values for col in df_metric.columns]
        
        try:
            stat, p_value = friedmanchisquare(*data_list)
            friedman_pvalues[metric] = p_value
            
            # Find the best model for this metric (using median)
            medians = {model: np.median(fold_results[model][metric]) for model in model_groups}
            best_model = max(medians, key=medians.get)
            
            # Post-hoc test: check if best model is significantly better than baseline
            best_significant_vs_baseline[metric] = False
            if p_value < 0.05:
                nemenyi_results = sp.posthoc_nemenyi_friedman(df_metric)
                print(f"  {metric}: Friedman p={p_value:.4f}, best_model={best_model}")
                # Only check if best model is NOT baseline
                if best_model != "baseline":
                    p_val_vs_baseline = nemenyi_results.loc[best_model, "baseline"]
                    best_significant_vs_baseline[metric] = p_val_vs_baseline < 0.05
                    print(f"    -> {best_model} vs baseline: p={p_val_vs_baseline:.4f}, significant={p_val_vs_baseline < 0.05}")
                else:
                    print(f"    -> baseline is best, no asterisk needed")
            else:
                print(f"  {metric}: Friedman p={p_value:.4f} (not significant, no post-hoc)")
        except Exception as e:
            print(f"  Warning: Friedman test failed for {metric}: {e}")
            friedman_pvalues[metric] = np.nan
            best_significant_vs_baseline[metric] = False
    
    return metrics_summary, friedman_pvalues, best_significant_vs_baseline


# ============== MAIN EXECUTION ==============
def main():
    # Define configurations
    # Configuration 1: [25, 25, 25, 1] - Extra Features
    config1 = {
        "name": "[25, 25, 25, 1]",
        "processed_file": "inflated_models/large_layers_overtrained_peptide.pkl",
        "splits_file": "large_layers_cv_splits/large_layers_cv_splits_peptide.pkl",
        "model_patterns": {
            "baseline": "extra_features_baseline_peptide/baseline_peptide_fold_*.h5",
            "method1": "transfer_learning_extra_features_smt_to_p/freeze_gnn_small_mol_tox_to_peptide_fold_*.h5",
            "method2": "transfer_learning_extra_features_smt_to_p/freeze_readout_small_mol_tox_to_peptide_fold_*.h5",
            "method3": "transfer_learning_extra_features_smt_to_p/freeze_all_small_mol_tox_to_peptide_fold_*.h5",
            "method4": "transfer_learning_extra_features_smt_to_p/gradual_unfreezing_small_mol_tox_to_peptide_fold_*.h5",
        }
    }
    
    # Configuration 2: [125, 125, 125, 1] - Large Layers
    config2 = {
        "name": "[125, 125, 125, 1]",
        "processed_file": "large_layers_models/large_layers_overtrained_peptide.pkl",
        "splits_file": "large_layers_cv_splits/large_layers_cv_splits_peptide.pkl",
        "model_patterns": {
            "baseline": "large_layers_baseline_peptide/large_layers_baseline_peptide_fold_*.h5",
            "method1": "large_layers_transfer_learning_smt_to_p/LL_freeze_gnn_smile_mol_tox_to_peptide_fold_*.h5",
            "method2": "large_layers_transfer_learning_smt_to_p/LL_freeze_readout_smile_mol_tox_to_peptide_fold_*.h5",
            "method3": "large_layers_transfer_learning_smt_to_p/LL_freeze_all_smile_mol_tox_to_peptide_fold_*.h5",
            "method4": "large_layers_transfer_learning_smt_to_p/LL_gradual_unfreezing_smile_mol_tox_to_peptide_fold_*.h5",
        }
    }
    
    # Configuration 3: [512, 256, 128, 1] - Inflated
    config3 = {
        "name": "[512, 256, 128, 1]",
        "processed_file": "inflated_models/large_layers_overtrained_peptide.pkl",
        "splits_file": "large_layers_cv_splits/large_layers_cv_splits_peptide.pkl",
        "model_patterns": {
            "baseline": "inflated_baseline_peptide/inflated_baseline_peptide_fold_*.h5",
            "method1": "inflated_transfer_learning_smt_to_p/inflated_freeze_gnn_smile_mol_tox_to_peptide_fold_*.h5",
            "method2": "inflated_transfer_learning_smt_to_p/inflated_freeze_readout_smile_mol_tox_to_peptide_fold_*.h5",
            "method3": "inflated_transfer_learning_smt_to_p/inflated_freeze_all_smile_mol_tox_to_peptide_fold_*.h5",
            "method4": "inflated_transfer_learning_smt_to_p/inflated_gradual_unfreezing_smile_mol_tox_to_peptide_fold_*.h5",
        }
    }
    
    configurations = [config1, config2, config3]
    
    # Model display names
    model_display_names = {
        "baseline": "Baseline",
        "method1": "Method 1 - freeze gnn",
        "method2": "Method 2 - freeze readout",
        "method3": "Method 3 - freeze all",
        "method4": "Method 4 - gradual unfreezing",
    }
    
    metric_names = ["ROC_AUC", "GM", "Precision", "Recall", "F1", "MCC"]
    metric_display_names = {
        "ROC_AUC": "ROC-AUC",
        "GM": "GM",
        "Precision": "Precision",
        "Recall": "Recall",
        "F1": "F1",
        "MCC": "MCC"
    }
    
    # Collect all results
    all_results = {}
    
    for config in configurations:
        metrics_summary, friedman_pvalues, significant_vs_baseline = evaluate_configuration(
            config["name"],
            config["processed_file"],
            config["splits_file"],
            config["model_patterns"]
        )
        all_results[config["name"]] = {
            "metrics_summary": metrics_summary,
            "friedman_pvalues": friedman_pvalues,
            "significant_vs_baseline": significant_vs_baseline
        }
    
    # ============== CREATE EXCEL FILE ==============
    print("\n" + "="*60)
    print("Creating Excel file (MEDIAN)...")
    print("="*60)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Peptide Evaluation (Median)"
    
    # Headers
    headers = ["Veličine GNN slojeva", "Modeli", "ROC-AUC", "GM", "Precision", "Recall", "F1", "MCC"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Populate data
    current_row = 2
    model_groups = ["baseline", "method1", "method2", "method3", "method4"]
    
    for config in configurations:
        config_name = config["name"]
        results = all_results[config_name]
        metrics_summary = results["metrics_summary"]
        friedman_pvalues = results["friedman_pvalues"]
        best_significant_vs_baseline = results["significant_vs_baseline"]
        
        # Find best values for each metric in this configuration (using median)
        best_values = {}
        best_models = {}
        for metric in metric_names:
            best_val = -np.inf
            best_model = None
            for group in model_groups:
                val = metrics_summary[group][metric]["median"]
                if not np.isnan(val) and val > best_val:
                    best_val = val
                    best_model = group
            best_values[metric] = best_val
            best_models[metric] = best_model
        
        # Write rows for each model
        for i, group in enumerate(model_groups):
            # GNN layer size (only on first row of group)
            if i == 0:
                ws.cell(row=current_row, column=1, value=config_name)
            
            # Model name
            ws.cell(row=current_row, column=2, value=model_display_names[group])
            
            # Metrics
            for col, metric in enumerate(metric_names, start=3):
                median_val = metrics_summary[group][metric]["median"]
                std_val = metrics_summary[group][metric]["std"]
                
                if np.isnan(median_val):
                    cell_value = "N/A"
                else:
                    cell_value = f"{median_val:.4f} ± {std_val:.4f}"
                
                cell = ws.cell(row=current_row, column=col, value=cell_value)
                
                # Check if this is the best value
                is_best = (best_models[metric] == group)
                
                # Apply formatting: Bold for best, asterisk only if best AND significantly different from baseline
                if is_best and not np.isnan(median_val):
                    # Asterisk only if: this is the best model AND it's significantly better than baseline
                    # (best_significant_vs_baseline is True only when best model != baseline and p < 0.05)
                    if best_significant_vs_baseline.get(metric, False):
                        cell.value = f"{median_val:.4f} ± {std_val:.4f} *"
                        cell.font = Font(bold=True)
                    else:
                        cell.font = Font(bold=True)
            
            current_row += 1
        
        # Friedman p-value row
        ws.cell(row=current_row, column=2, value="Friedman p-value")
        for col, metric in enumerate(metric_names, start=3):
            p_val = friedman_pvalues.get(metric, np.nan)
            if np.isnan(p_val):
                cell_value = "N/A"
            elif p_val < 0.0001:
                # Use scientific notation for very small p-values
                cell_value = f"{p_val:.2e}"
            else:
                cell_value = f"{p_val:.4f}"
            ws.cell(row=current_row, column=col, value=cell_value)
        
        current_row += 1
    
    # Adjust column widths
    column_widths = [20, 35, 20, 20, 20, 20, 20, 20]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save workbook
    output_file = "eval-book-peptide-results-median.xlsx"
    wb.save(output_file)
    print(f"\nResults saved to: {output_file}")
    
    # Also print summary to console
    print("\n" + "="*60)
    print("SUMMARY (MEDIAN)")
    print("="*60)
    for config in configurations:
        config_name = config["name"]
        print(f"\n{config_name}:")
        metrics_summary = all_results[config_name]["metrics_summary"]
        for group in model_groups:
            print(f"  {model_display_names[group]}:")
            for metric in metric_names:
                median_val = metrics_summary[group][metric]["median"]
                std_val = metrics_summary[group][metric]["std"]
                print(f"    {metric}: {median_val:.4f} ± {std_val:.4f}")


if __name__ == "__main__":
    main()
